import openpyxl
import io


async def extract_text_from_xlsx(stream: io.BytesIO) -> str:
    """
    Extracts text-based data from an .xlsx file using openpyxl.
    Iterates through all sheets, rows, and cells, and concatenates
    non-empty cell values into a readable string.

    Args: stream (io.BytesIO): In-memory XLSX file stream

    Returns: str: Cleaned and formatted text from the workbook
    """
    try:
        workbook = openpyxl.load_workbook(stream, data_only=True)
        extracted_lines = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            extracted_lines.append(f"--- Sheet: {sheet_name} ---")

            for row in sheet.iter_rows(values_only=True):
                # Filter out None values and convert to string
                non_empty_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if non_empty_values:
                    extracted_lines.append("\t\t".join(non_empty_values))

            extracted_lines.append("")  # Spacer between sheets

        workbook.close()
        return "\n\n".join(extracted_lines)

    except Exception as e:
        raise ValueError(f"Failed to extract text from XLSX: {e}")
